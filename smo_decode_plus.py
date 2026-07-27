#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""Advanced ScanMyOpel *.smo log decoder.

Decodes .smo files and produces structured reports with ECU info, DTC codes,
and live data tables. Supports plain-text (.txt) and Excel (.xlsx) output.
In Excel mode Live Data is transposed (rows = cycles, columns = params) and
includes monochrome line charts with tight Y-axis scaling.
"""

from pathlib import Path
import sys
import subprocess
from typing import Dict, List, Optional, Tuple

__version__ = "0.2.13"  # Build counter: increment last digit on every release

# -----------------------------------------------------------------------------
# Localization
# -----------------------------------------------------------------------------
_MESSAGES: Dict[str, Dict[str, str]] = {
    "en": {
        "warning_odd_length": "Warning: odd file length.",
        "file_not_found": "File not found:",
        "no_smo_files": "No .smo files found in directory '{}'.",
        "files_found": "Files found: {}",
        "done": "Done.",
        "result_file": "Result saved to:",
        "processed_ok": "Successfully processed: {}",
        "errors": "Errors: {}",
        "error_processing": "Error processing '{}':",
        "file_too_large": "File is too large: {:.1f} MB (max {:.1f} MB)",
        "write_error": "Failed to write output file",
        "invalid_path": "Invalid file path: {}",
        "ecu_info": "ECU INFORMATION",
        "dtc_codes": "DIAGNOSTIC TROUBLE CODES (DTC)",
        "no_errors": "No errors found.",
        "live_data": "LIVE DATA",
        "total_cycles": "Total polling cycles recorded: {}",
        "no_live_data": "Live Data not detected.",
        "value_table": "VALUE TABLE BY CYCLE",
        "parameter": "Parameter",
        "changes_graphs": "CHANGE GRAPHS",
        "min": "Min",
        "max": "Max",
        "cleared": "--- Trouble Codes Cleared ---",
        "xlsx_missing": "openpyxl is required for Excel output.",
        "xlsx_prompt": "Install openpyxl now? [Y/n]: ",
        "xlsx_install_ok": "openpyxl installed successfully.",
        "xlsx_install_fail": "Failed to install openpyxl: {}",
        "usage": "Usage: {} [--xlsx|--txt] [-l ru|en] [file.smo]",
    },
    "ru": {
        "warning_odd_length": "Предупреждение: нечётная длина файла.",
        "file_not_found": "Файл не найден:",
        "no_smo_files": "В папке '{}' файлы .smo не найдены.",
        "files_found": "Найдено файлов: {}",
        "done": "Готово.",
        "result_file": "Результат сохранён в:",
        "processed_ok": "Успешно обработано: {}",
        "errors": "Ошибок: {}",
        "error_processing": "Ошибка при обработке '{}':",
        "file_too_large": "Файл слишком большой: {:.1f} МБ (макс. {:.1f} МБ)",
        "write_error": "Не удалось записать выходной файл",
        "invalid_path": "Некорректный путь к файлу: {}",
        "ecu_info": "ИНФОРМАЦИЯ ОБ ЭБУ",
        "dtc_codes": "КОДЫ ОШИБОК (DTC)",
        "no_errors": "Ошибок не найдено.",
        "live_data": "ЖИВЫЕ ДАННЫЕ",
        "total_cycles": "Всего зафиксировано тактов опроса: {}",
        "no_live_data": "Данные Live Data не обнаружены.",
        "value_table": "ТАБЛИЦА ЗНАЧЕНИЙ ПО ТАКТАМ",
        "parameter": "Параметр",
        "changes_graphs": "ГРАФИКИ ИЗМЕНЕНИЙ",
        "min": "Min",
        "max": "Max",
        "cleared": "--- Ошибки стёрты (Trouble Codes Cleared) ---",
        "xlsx_missing": "Для Excel-вывода нужен openpyxl.",
        "xlsx_prompt": "Установить openpyxl сейчас? [Y/n]: ",
        "xlsx_install_ok": "openpyxl успешно установлен.",
        "xlsx_install_fail": "Не удалось установить openpyxl: {}",
        "usage": "Использование: {} [--xlsx|--txt] [-l ru|en] [file.smo]",
    },
}


def _get_messages(lang: str = "en") -> Dict[str, str]:
    """Return localized message dictionary."""
    return _MESSAGES.get(lang, _MESSAGES["en"])


# -----------------------------------------------------------------------------
# Constants
# -----------------------------------------------------------------------------
MAX_FILE_SIZE: int = 50 * 1024 * 1024  # 50 MB safety limit
SPARKLINE_CHUNK: int = 80  # Max characters per sparkline segment (rev. point #7)
COL_WIDTH: int = 8  # Table column width for cycle values
A4_MAX_WIDTH: int = 60  # Excel column width units, approx A4 portrait with margins


# -----------------------------------------------------------------------------
# Optional Excel dependency handling
# -----------------------------------------------------------------------------
def ensure_openpyxl(messages: Dict[str, str]) -> None:
    """Ensure openpyxl is available, prompt user to install dynamically if missing.

    Args:
        messages: Localized message dictionary.

    Raises:
        RuntimeError: If openpyxl is missing and user declines or install fails.
    """
    try:
        import importlib

        importlib.import_module("openpyxl")
        return
    except ImportError:
        print(messages["xlsx_missing"])
        try:
            choice = input(messages["xlsx_prompt"]).strip().lower()
        except (EOFError, KeyboardInterrupt):
            choice = "n"
        if choice in ("", "y", "yes"):
            try:
                subprocess.check_call(
                    [sys.executable, "-m", "pip", "install", "openpyxl"]
                )
                print(messages["xlsx_install_ok"])
                # Force re-import to pick up the newly installed package
                if "openpyxl" in sys.modules:
                    del sys.modules["openpyxl"]
                import importlib

                importlib.import_module("openpyxl")
                return
            except Exception as exc:
                raise RuntimeError(messages["xlsx_install_fail"].format(exc)) from exc
        raise RuntimeError(messages["xlsx_missing"])


# -----------------------------------------------------------------------------
# Core decoder
# -----------------------------------------------------------------------------
def decode_smo(data: bytes) -> bytes:
    """Decode ScanMyOpel *.smo file content.

    The encoding scheme maps each byte to two ASCII characters (A-P range),
    applies XOR 0x05 and nibble swap. Newline marker is 0x5F ('KF').

    Args:
        data: Raw encoded bytes from .smo file.

    Returns:
        Decoded bytearray as bytes.
    """
    out = bytearray()

    if len(data) % 2 != 0:
        print(_get_messages()["warning_odd_length"])

    for i in range(0, len(data) - 1, 2):
        c1 = data[i]
        c2 = data[i + 1]

        # Original encoding logic preserved (see rev. point #13)
        lo = ((c1 - ord("A") + 10) ^ 5) & 0x0F
        hi = ((c2 - ord("A") + 10) ^ 5) & 0x0F

        value = lo | (hi << 4)

        if value == 0x5F:
            out.append(ord("\n"))
            continue

        out.append(value)

    return bytes(out)


# -----------------------------------------------------------------------------
# Input validation and I/O helpers
# -----------------------------------------------------------------------------
def validate_path(path: Path) -> Path:
    """Validate that the path does not escape the working directory.

    Args:
        path: Input file path.

    Returns:
        Resolved absolute path.

    Raises:
        ValueError: If path points outside the current working directory.
    """
    resolved = path.resolve()
    cwd = Path.cwd().resolve()
    try:
        resolved.relative_to(cwd)
    except ValueError as exc:
        raise ValueError(_get_messages()["invalid_path"].format(path)) from exc
    return resolved


def check_file_size(path: Path, max_size: int = MAX_FILE_SIZE) -> None:
    """Check that the file size is within the allowed limit.

    Args:
        path: File to check.
        max_size: Maximum allowed size in bytes.

    Raises:
        ValueError: If the file exceeds max_size.
    """
    size = path.stat().st_size
    if size > max_size:
        raise ValueError(
            _get_messages()["file_too_large"].format(
                size / 1024 / 1024, max_size / 1024 / 1024
            )
        )


def safe_write(path: Path, content: str, messages: Dict[str, str]) -> None:
    """Write text to file with error handling.

    Args:
        path: Output file path.
        content: Text content to write.
        messages: Localized message dictionary for error strings.

    Raises:
        OSError: If writing fails after logging the error.
    """
    try:
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)
    except OSError as exc:
        print(f"{messages['write_error']}: {path} — {exc}")
        raise


# -----------------------------------------------------------------------------
# Log parsing helpers
# -----------------------------------------------------------------------------
def is_hex_garbage(line: str) -> bool:
    """Check if a line is raw HEX dump.

    Note: The original implementation is preserved for compatibility
    with the observed encoding behavior (rev. point #1).
    """
    parts = line.split()
    if len(parts) < 3:
        return False
    return all(p in "0123456789ABCDEF" for p in parts)


def parse_log(
    text: str, skip_patterns: Tuple[str, ...], messages: Dict[str, str]
) -> Tuple[List[str], List[str], List[Dict[str, str]]]:
    """Parse decoded log text into ECU info, DTCs, and live data cycles.

    Args:
        text: Decoded log text.
        skip_patterns: Tuple of substring patterns to skip (rev. point #2).
        messages: Localized message dictionary.

    Returns:
        Tuple of (info_lines, dtc_lines, cycles).
    """
    lines = text.splitlines()
    info_lines: List[str] = []
    dtc_lines: List[str] = []
    cycles: List[Dict[str, str]] = []

    current_cycle: Dict[str, str] = {}
    in_param_name: bool = False
    param_name_parts: List[str] = []

    last_param_name: str = ""
    param_has_string_value: bool = False
    last_param_value: str = ""

    for line in lines:
        line = line.strip()
        if not line:
            continue

        if is_hex_garbage(line):
            continue

        # Original substring matching preserved (rev. point #2)
        if any(skip in line for skip in skip_patterns):
            if "Data array Length is 0" in line and dtc_lines:
                dtc_lines.append(messages["cleared"])
            continue

        if line.startswith("AT") or line in (">", ">>"):
            continue

        # ECU static info
        if "Name:" in line and "Value:" in line:
            info_lines.append(line.split(" - ", 1)[-1])
            continue

        # DTC codes
        if "Code:" in line and "Name:" in line:
            err = line.split(" - ", 1)[-1]
            if err not in dtc_lines:
                dtc_lines.append(err)
            continue

        # Live Data cycle assembly
        if "Parameter Name -" in line:
            if last_param_name:
                current_cycle[last_param_name] = last_param_value

            in_param_name = True
            param_name_parts = []
            last_param_value = ""
            param_has_string_value = False
            last_param_name = ""
            continue

        if in_param_name:
            if "Parameter Formula" in line:
                in_param_name = False
                last_param_name = " ".join(param_name_parts).strip()
            else:
                clean_part = line.split(" - ", 1)[-1] if " - " in line else line
                param_name_parts.append(clean_part)
            continue

        if "Value S -" in line:
            val = line.split("Value S -", 1)[-1].strip()
            if val != "null":
                last_param_value = val
                param_has_string_value = True
            continue

        if "Value D -" in line:
            if not param_has_string_value:
                last_param_value = line.split("Value D -", 1)[-1].strip()
            continue

        # End of polling cycle
        if "Update adapter info" in line:
            if last_param_name:
                current_cycle[last_param_name] = last_param_value
                last_param_name = ""

            if current_cycle:
                cycles.append(current_cycle)
                current_cycle = {}
            continue

    # Capture trailing parameter and cycle
    if last_param_name:
        current_cycle[last_param_name] = last_param_value
    if current_cycle:
        cycles.append(current_cycle)

    return info_lines, dtc_lines, cycles


# -----------------------------------------------------------------------------
# TXT report builders
# -----------------------------------------------------------------------------
def build_table(
    cycles: List[Dict[str, str]],
    all_params: List[str],
    messages: Dict[str, str],
) -> List[str]:
    """Build ASCII table of parameter values per polling cycle.

    Args:
        cycles: List of parameter dictionaries per cycle.
        all_params: Ordered list of unique parameter names.
        messages: Localized message dictionary.

    Returns:
        List of table lines.
    """
    report: List[str] = []
    report.append(f"=== {messages['value_table']} ===")

    header = f"{messages['parameter']:<35} | " + " | ".join(
        [f"T{i}".ljust(COL_WIDTH) for i in range(len(cycles))]
    )
    report.append(header)
    report.append("-" * len(header))

    for p in all_params:
        row_vals: List[str] = []
        for cycle in cycles:
            val = cycle.get(p, "-")
            row_vals.append(_format_table_value(val))

        p_name = (p[:33] + "..") if len(p) > 33 else p
        report.append(f"{p_name:<35} | " + " | ".join(row_vals))

    return report


def _format_table_value(val: Optional[str], width: int = COL_WIDTH) -> str:
    """Format a single value for the ASCII table.

    Fixes rev. point #8: uses fixed-width rounding instead of hard truncation.

    Args:
        val: Raw value (string, number, or None).
        width: Target column width.

    Returns:
        Left-justified string fitting within width.
    """
    if val is None or val == "-":
        return "-".ljust(width)
    try:
        num = float(val)
        formatted = f"{num:.2f}"
        if len(formatted) > width:
            formatted = f"{num:.1f}"
        return formatted[:width].ljust(width)
    except (ValueError, TypeError):
        return str(val)[:width].ljust(width)


def build_sparklines(
    cycles: List[Dict[str, str]],
    all_params: List[str],
    messages: Dict[str, str],
    chunk_size: int = SPARKLINE_CHUNK,
) -> List[str]:
    """Build sparkline graphs for numeric parameters, chunked by line length.

    Implements rev. point #7: long sparklines are split across multiple lines.
    Filters out parameters whose values do not change (rev. point #16).
    Adds tick ruler (every 5th cycle) under each chunk (rev. point #17).

    Args:
        cycles: List of parameter dictionaries per cycle.
        all_params: Ordered list of unique parameter names.
        messages: Localized message dictionary.
        chunk_size: Maximum characters per sparkline segment.

    Returns:
        List of sparkline report lines.
    """
    report: List[str] = []
    report.append(f"\n=== {messages['changes_graphs']} ===")
    spark_chars = "▁▂▃▄▅▆▇█"

    for p in all_params:
        values: List[Optional[float]] = []
        is_numeric = True

        for cycle in cycles:
            val = cycle.get(p)
            if val is None or val == "-":
                values.append(None)
            else:
                try:
                    values.append(float(val))
                except ValueError:
                    is_numeric = False
                    break

        if not is_numeric:
            continue

        valid_vals = [v for v in values if v is not None]
        if not valid_vals:
            continue

        min_v = min(valid_vals)
        max_v = max(valid_vals)

        # Фильтр: пропускаем параметры, значения которых не меняются (rev. #16)
        if min_v == max_v:
            continue

        range_v = max_v - min_v

        # Build full sparkline
        full_spark = ""
        for v in values:
            if v is None:
                full_spark += " "
            else:
                norm = (v - min_v) / range_v
                idx = int(norm * 7)
                idx = max(0, min(7, idx))
                full_spark += spark_chars[idx]

        # Compact numeric formatting (rev. #15)
        min_str = f"{min_v:g}"
        max_str = f"{max_v:g}"

        report.append(f"{p}")
        report.append(
            f"  {messages['min']}: {min_str:<10} {messages['max']}: {max_str:<10}"
        )

        # Chunk the sparkline + tick ruler under each chunk (rev. #7, #17)
        total_ticks = len(full_spark)
        for chunk_start in range(0, total_ticks, chunk_size):
            chunk_end = min(chunk_start + chunk_size, total_ticks)
            segment = full_spark[chunk_start:chunk_end]
            seg_len = len(segment)

            # Линейка с метками каждого 5-го такта (без буквы T)
            ruler = [" "] * seg_len
            for i in range(seg_len):
                tick = chunk_start + i
                if tick % 5 == 0:
                    label = f"{tick}"
                    for j, ch in enumerate(label):
                        pos = i + j
                        if pos < seg_len:
                            ruler[pos] = ch
            ruler_line = "".join(ruler).rstrip()

            report.append(f"  | {segment}")
            if ruler_line:
                report.append(f"  | {ruler_line}")
        report.append("")

    return report


def build_txt_report(
    info_lines: List[str],
    dtc_lines: List[str],
    cycles: List[Dict[str, str]],
    messages: Dict[str, str],
) -> str:
    """Build the full plain-text report.

    Args:
        info_lines: ECU static information lines.
        dtc_lines: Diagnostic trouble code lines.
        cycles: Live data polling cycles.
        messages: Localized message dictionary.

    Returns:
        Complete report as a single string.
    """
    report: List[str] = []
    report.append(f"=== {messages['ecu_info']} ===")
    report.extend(info_lines)
    report.append(f"\n=== {messages['dtc_codes']} ===")
    if dtc_lines:
        report.extend(dtc_lines)
    else:
        report.append(messages["no_errors"])

    report.append(f"\n=== {messages['live_data']} ===")
    report.append(f"{messages['total_cycles'].format(len(cycles))}\n")

    if not cycles:
        report.append(messages["no_live_data"])
        return "\n".join(report)

    # Rev. point #6: O(1) ordered unique params via dict keys (Python 3.7+)
    all_params = list(dict.fromkeys(p for cycle in cycles for p in cycle))

    report.extend(build_table(cycles, all_params, messages))
    report.extend(build_sparklines(cycles, all_params, messages))

    return "\n".join(report)


# -----------------------------------------------------------------------------
# XLSX helpers
# -----------------------------------------------------------------------------
def _to_numeric_or_str(val: Optional[str]):
    """Convert a string value to float/int if possible, else return as-is.

    Args:
        val: Raw string value or None.

    Returns:
        float, int, str, or empty string.
    """
    if val is None or val == "-" or val == "":
        return ""
    try:
        f = float(val)
        if f.is_integer():
            return int(f)
        return f
    except (ValueError, TypeError):
        return str(val)


def _set_a4_col_width(ws, col_idx: int, content_width: int) -> None:
    """Set column width capped at A4 portrait limit; enable wrap if truncated.

    Args:
        ws: openpyxl worksheet.
        col_idx: 1-based column index.
        content_width: Desired width based on content.
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment

    letter = get_column_letter(col_idx)
    if content_width > A4_MAX_WIDTH:
        ws.column_dimensions[letter].width = A4_MAX_WIDTH
        # Enable wrap for all existing cells in this column
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if not cell.alignment.wrap_text:
                    cell.alignment = Alignment(
                        wrap_text=True, vertical="top", horizontal="left"
                    )
    else:
        ws.column_dimensions[letter].width = max(content_width, 8)


def _style_header_row(ws, cols: int) -> None:
    """Apply bold font, light fill and bottom border to the first row.

    Args:
        ws: openpyxl worksheet.
        cols: Number of columns to style.
    """
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"
    )
    thin_border = Border(bottom=Side(style="thin", color="000000"))
    for col in range(1, cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        # Top alignment so wrapped headers start from the top edge
        cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)


# -----------------------------------------------------------------------------
# XLSX report builder (transposed: rows = cycles, columns = params)
# -----------------------------------------------------------------------------
def build_xlsx_report(
    info_lines: List[str],
    dtc_lines: List[str],
    cycles: List[Dict[str, str]],
    output_path: Path,
    messages: Dict[str, str],
) -> None:
    """Build a transposed Excel report with monochrome charts.

    Sheets:
        - ECU Info: static ECU data with A4-safe column widths.
        - DTC: diagnostic trouble codes with A4-safe column widths.
        - Live Data: transposed table where each row is a polling cycle (T0, T1…)
          and each column is a parameter. Column widths are based on data values,
          not headers; truncated headers get Excel comments with full text.
        - Charts: monochrome line charts with tight Y-axis scaling.

    Args:
        info_lines: ECU static information lines.
        dtc_lines: Diagnostic trouble code lines.
        cycles: Live data polling cycles.
        output_path: Destination .xlsx file path.
        messages: Localized message dictionary.
    """
    ensure_openpyxl(messages)

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.marker import Marker
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.line import LineProperties
    from openpyxl.comments import Comment

    wb = openpyxl.Workbook()

    # --- Sheet 1: ECU Info ---
    ws_info = wb.active
    ws_info.title = "ECU Info"
    ws_info.append(["Key", "Value"])
    max_key = 10
    max_val = 10
    for line in info_lines:
        if "Value:" in line:
            key, _, value = line.partition(" Value: ")
            key_clean = key.replace("Name: ", "").strip()
            ws_info.append([key_clean, value.strip()])
            max_key = max(max_key, len(key_clean))
            max_val = max(max_val, len(value.strip()))
        else:
            ws_info.append([line, ""])
            max_key = max(max_key, len(line))
    _style_header_row(ws_info, 2)
    _set_a4_col_width(ws_info, 1, max_key + 2)
    _set_a4_col_width(ws_info, 2, max_val + 2)
    ws_info.freeze_panes = "A2"

    # --- Sheet 2: DTC ---
    ws_dtc = wb.create_sheet(title="DTC")
    ws_dtc.append([messages["dtc_codes"]])
    max_len = len(messages["dtc_codes"])
    if dtc_lines:
        for line in dtc_lines:
            ws_dtc.append([line])
            max_len = max(max_len, len(line))
    else:
        ws_dtc.append([messages["no_errors"]])
    _style_header_row(ws_dtc, 1)
    _set_a4_col_width(ws_dtc, 1, max_len + 2)
    ws_dtc.freeze_panes = "A2"

    # --- Sheet 3: Live Data (transposed) ---
    ws_live = wb.create_sheet(title="Live Data")
    all_params = list(dict.fromkeys(p for cycle in cycles for p in cycle))

    # Header row: Cycle | Param1 | Param2 | ...
    header = [messages["parameter"]] + all_params
    ws_live.append(header)

    # Set header row height for 2-line wrapping with top alignment
    ws_live.row_dimensions[1].height = 30

    # Data rows: 0 | val1 | val2 | ... (без буквы T для чистой оси графиков)
    for i, cycle in enumerate(cycles):
        row = [i]
        for p in all_params:
            val = cycle.get(p, "")
            row.append(_to_numeric_or_str(val))
        ws_live.append(row)

    _style_header_row(ws_live, len(header))

    # Freeze panes: row 1 + column A
    ws_live.freeze_panes = "B2"

    # Column widths based on data values (not headers), add comments for long headers
    # Col 1 = cycle labels
    _set_a4_col_width(ws_live, 1, max(6, len(str(len(cycles))) + 2))

    for col_idx, param in enumerate(all_params, start=2):
        # Compute max width among data values (skip header)
        max_val_len = 0
        for cycle in cycles:
            val = cycle.get(param, "")
            val_str = str(val)[:30]  # safety cap
            max_val_len = max(max_val_len, len(val_str))
        # Minimum 8, cap at A4
        col_width = max(8, min(max_val_len + 2, A4_MAX_WIDTH))
        _set_a4_col_width(ws_live, col_idx, col_width)

        # Add comment with full parameter name if header would be truncated
        header_cell = ws_live.cell(row=1, column=col_idx)
        if len(param) > col_width:
            header_cell.comment = Comment(param, "smo_decode_plus")

    # --- Sheet 4: Charts ---
    ws_charts = wb.create_sheet(title="Charts")
    chart_idx = 0

    for col_idx, param in enumerate(all_params, start=2):
        # Determine if parameter is numeric and changing
        values: List[Optional[float]] = []
        is_numeric = True
        for cycle in cycles:
            val = cycle.get(param)
            if val is None or val == "":
                values.append(None)
            else:
                try:
                    values.append(float(val))
                except (ValueError, TypeError):
                    is_numeric = False
                    break

        if not is_numeric:
            continue
        valid = [v for v in values if v is not None]
        if len(valid) < 2 or max(valid) == min(valid):
            continue

        min_v = min(valid)
        max_v = max(valid)

        # Build line chart
        chart = LineChart()
        chart.title = param[:31]
        chart.style = 2
        chart.y_axis.title = None
        chart.x_axis.title = None
        chart.height = 10
        chart.width = 18

        data_ref = Reference(
            ws_live, min_col=col_idx, min_row=1, max_row=len(cycles) + 1
        )
        cats_ref = Reference(ws_live, min_col=1, min_row=2, max_row=len(cycles) + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        # Remove legend
        chart.legend = None

        # 1. Вспомогательная функция для расчёта цвета (зелёный -> оранжевый -> красный)
        def _get_gradient_color(val: float, min_v: float, max_v: float) -> str:
            if max_v == min_v:
                return "FFC000"  # Оранжевый, если значения не меняются
            ratio = (val - min_v) / (max_v - min_v)
            if ratio < 0.5:
                # От зелёного (00B050) к оранжевому (FFC000)
                r = ratio * 2
                R = round(0 + r * 255)
                G = round(176 + r * (192 - 176))
                B = round(80 + r * (0 - 80))
            else:
                # От оранжевого (FFC000) к красному (FF0000)
                r = (ratio - 0.5) * 2
                R = 255
                G = round(192 + r * (0 - 192))
                B = 0
            return f"{R:02X}{G:02X}{B:02X}"

        # Импортируем DataPoint прямо здесь, чтобы избежать ошибок
        from openpyxl.chart.series import DataPoint

        # 2. Базовые настройки (серая линия, на случай пропусков данных)
        s = chart.series[0]
        s.graphicalProperties.line.solidFill = "BFBFBF"  # Серая базовая линия
        s.graphicalProperties.line.w = 40000  # Жирная линия (w вместо width)

        s.marker = Marker(symbol="circle", size=7)
        s.marker.graphicalProperties = GraphicalProperties(solidFill="BFBFBF")
        s.marker.graphicalProperties.line = LineProperties(solidFill="BFBFBF")

        # 3. Индивидуальная раскраска каждого такта (точки и отрезка линии)
        for idx, v in enumerate(values):
            if v is None:
                continue
            color = _get_gradient_color(v, min_v, max_v)
            dp = DataPoint(idx=idx)

            # Цвет и размер точки (узла)
            dp.marker = Marker(symbol="circle", size=7)
            dp.marker.graphicalProperties = GraphicalProperties(solidFill=color)
            dp.marker.graphicalProperties.line = LineProperties(solidFill=color)

            # Цвет отрезка линии, ведущего к этой точке
            dp.graphicalProperties = GraphicalProperties()
            dp.graphicalProperties.line = LineProperties(solidFill=color, w=40000)
            s.data_points.append(dp)

        # X-axis: show every 5th tick label (rev. #9)
        chart.x_axis.delete = False
        chart.x_axis.tickLblSkip = 5
        chart.x_axis.tickLblPos = "low"

        # Y-axis: tight scaling with small padding
        pad = max((max_v - min_v) * 0.15, 0.5)
        chart.y_axis.scaling.min = min_v - pad
        chart.y_axis.scaling.max = max_v + pad

        # Show key values (min / max / intermediate) on Y axis (rev. #18)
        from openpyxl.chart.axis import ChartLines

        chart.y_axis.delete = False
        chart.y_axis.tickLblPos = "low"
        chart.y_axis.numFmt = "0.##"
        chart.y_axis.majorGridlines = ChartLines()  # включаем сетку с подписями
        chart.y_axis.minorGridlines = None

        # Position charts vertically
        row_anchor = chart_idx * 20 + 1
        anchor = f"A{row_anchor}"
        ws_charts.add_chart(chart, anchor)

        chart_idx += 1

    wb.save(str(output_path))


# -----------------------------------------------------------------------------
# File conversion
# -----------------------------------------------------------------------------
def convert_file(path: Path, lang: str = "en", use_xlsx: bool = False) -> None:
    """Convert a single .smo file to a report.

    Args:
        path: Path to the .smo file.
        lang: Language code ('en' or 'ru').
        use_xlsx: If True, write Excel output; otherwise plain text.
    """
    messages = _get_messages(lang)

    validated = validate_path(path)
    check_file_size(validated)

    with open(validated, "rb") as f:
        encoded = f.read()

    decoded = decode_smo(encoded)

    try:
        text = decoded.decode("utf-8")
    except UnicodeDecodeError:
        text = decoded.decode("utf-8", errors="replace")

    skip_patterns = (
        "Bluetooth State -",
        "Read block -",
        "Data block created!",
        "Message Group Sending",
        "****",
        "Abstract Activity",
        "Create Tabs",
        "Create -",
        "Android OS",
        "Device Name",
        "Device Model",
        "Package name",
        "App Version",
        "BUS INIT",
        "Repeat Data Reading",
        "Owner Changed",
        "Live Data Off",
        "Live Data On",
        "Adapter Test",
        "Adapter Init",
        "ECU Init",
        "Selector Init",
        "TC Read",
    )

    info_lines, dtc_lines, cycles = parse_log(text, skip_patterns, messages)

    if use_xlsx:
        output_path = path.with_suffix(".xlsx")
        build_xlsx_report(info_lines, dtc_lines, cycles, output_path, messages)
    else:
        report_text = build_txt_report(info_lines, dtc_lines, cycles, messages)
        output_path = path.with_suffix(".txt")
        safe_write(output_path, report_text, messages)

    print(f"{messages['done']} {path.name}")
    print(f"{messages['result_file']} {output_path.name}")


# -----------------------------------------------------------------------------
# CLI entry point
# -----------------------------------------------------------------------------
def main(argv: Optional[List[str]] = None) -> None:
    """Command-line entry point.

    Supports:
        --xlsx      Output to Excel (.xlsx) with transposed live data and charts.
        --txt       Output to plain text (.txt) [default].
        -l / --lang Language: 'en' or 'ru'.

    Args:
        argv: Optional argument list override (for testing).
    """
    if argv is None:
        argv = sys.argv[1:]

    lang = "en"
    use_xlsx = False
    args: List[str] = list(argv)

    # Parse flags
    if "--xlsx" in args:
        use_xlsx = True
        args.remove("--xlsx")

    if "--txt" in args:
        use_xlsx = False
        args.remove("--txt")

    if "--lang" in args:
        idx = args.index("--lang")
        if idx + 1 < len(args):
            lang = args[idx + 1]
            args.pop(idx)
            args.pop(idx)

    if "-l" in args:
        idx = args.index("-l")
        if idx + 1 < len(args):
            lang = args[idx + 1]
            args.pop(idx)
            args.pop(idx)

    if args:
        path = Path(args[0])
        if not path.exists():
            print(f"{_get_messages(lang)['file_not_found']} {path}")
            return
        convert_file(path, lang, use_xlsx)
        return

    current_dir = Path.cwd()
    files = sorted(current_dir.glob("*.smo"))

    if not files:
        print(_get_messages(lang)["no_smo_files"].format(current_dir))
        return

    print(f"{_get_messages(lang)['files_found'].format(len(files))}\n")

    ok = 0
    bad = 0

    for file in files:
        try:
            convert_file(file, lang, use_xlsx)
            ok += 1
        except Exception as exc:
            bad += 1
            print(f"\n{_get_messages(lang)['error_processing'].format(file)}")
            print(exc)

    messages = _get_messages(lang)
    print(f"\n{messages['done']}")
    print(f"{messages['processed_ok'].format(ok)}")
    print(f"{messages['errors'].format(bad)}")


if __name__ == "__main__":
    main()
